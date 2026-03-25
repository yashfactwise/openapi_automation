#!/usr/bin/env python3
"""
Generate the Excel template for bulk PO upload.
Run this once to create the .xlsx file with all columns and example data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Columns: (column_name, example_value, hint, required)
COLUMNS = [
    # ── Buyer ──
    ("Buyer Entity Name",           "FactWise",                                         "Your company/entity name",                     False),
    ("Buyer Billing Address",       "Main address",                                     "Billing address name or ID",                   False),
    ("Buyer Shipping Address",      "Second address",                                   "Shipping address name or ID",                  False),
    ("Buyer Identifications",       "GST",                                              "Comma-separated, e.g. GST,PAN",                False),
    ("Buyer Contacts",              "globalfielsETE@gmail.com",                         "Comma-separated phone/email",                  False),

    # ── Seller / Vendor ──
    ("Vendor ERP Code",             "",                                                 "Vendor code in your ERP system",               False),
    ("Vendor Factwise Code",        "FacVENDOR012620250600",                            "Factwise vendor code",                         False),
    ("Vendor Address ID",           "",                                                 "Vendor address UUID (optional)",               False),
    ("Vendor Address Text",         "66, Dwarkadas Building, Mangaldas Road, Kalbadevi, Mumbai 710700", "Vendor full address (free text)", False),
    ("Vendor ID Type",              "GST",                                              "Identification type, e.g. GST or PAN",         False),
    ("Vendor ID Value",             "wew23212",                                         "Identification number/value",                  False),
    ("Vendor Contacts",             "vendor0101@gmail.com",                             "Comma-separated emails",                       False),

    # ── PO Details ──
    ("Creator Email",               "globalfieldsETE@gmail.com",                        "Email of the user creating this PO",           False),
    ("ERP PO ID",                   "SS-ERP-35565",                                     "Unique PO ID from your ERP (must be unique)",  False),
    ("Status",                      "ISSUED",                                           "ISSUED or ONGOING",                            False),
    ("Template Name",               "PO Open API",                                      "PO template name in Factwise",                 False),
    ("Issued Date",                 "2025-02-24",                                       "YYYY-MM-DD",                                   False),
    ("Accepted Date",               "2026-05-29",                                       "YYYY-MM-DD",                                   False),
    ("Currency",                    "a8c3e3fd-b05f-4d09-bd2f-9fedd07d0ec3",             "Currency UUID",                                False),
    ("PO Notes",                    "",                                                 "General notes for this PO",                    False),
    ("Event",                       "",                                                 "Event name (optional)",                        False),
    ("TnC Name",                    "Ashir TnC",                                        "Terms & Conditions name",                      False),
    ("TnC Data",                    "These are my terms and these are my conditions. Do you agree or do you agree?", "T&C content",     False),
    ("Incoterm",                    "NA",                                               "Incoterm code, e.g. DAP, FOB, NA",             False),
    ("Prepayment %",                "0",                                                "0 to 100",                                     False),
    ("Payment Type",                "PER_INVOICE_ITEM",                                 "PER_INVOICE_ITEM or PER_DELIVERABLE",          False),
    ("Payment Terms",               '{"term": 2, "period": "MONTHS", "applied_from": "INVOICE_DATE"}', "JSON object",                  False),
    ("Deliverables Payment Terms",  "",                                                 "JSON array (optional)",                        False),
    ("Lead Time",                   "10",                                               "Number",                                       False),
    ("Lead Time Period",            "DAYS",                                             "DAYS, WEEKS, or MONTHS",                       False),
    ("PO Custom Sections",          "[]",                                               "JSON array — leave [] if no sections",         False),

    # ── Item ──
    ("Item ERP Code",               "AA-ERP-101",                                       "Item code in your ERP (optional)",             False),
    ("Item Factwise Code",          "AA-FW-01",                                         "Factwise item code",                           False),
    ("Item Additional Details",     "item level ad",                                    "Extra item description",                       False),
    ("Item Internal Notes",         "item level in",                                    "Internal notes (not visible to vendor)",       False),
    ("Item External Notes",         "item level en",                                    "External notes (visible to vendor)",           False),
    ("Item Price",                  "10",                                               "Unit price",                                   False),
    ("Item Quantity",               "10",                                               "Total quantity",                               False),
    ("Item Unit",                   "f16d124e-db59-48fe-a2b8-19f625745cbf",             "Measurement unit UUID",                        False),
    ("Item Incoterm",               "DAP",                                              "Item-level incoterm",                          False),
    ("Item Prepayment %",           "10",                                               "0 to 100",                                     False),
    ("Item Lead Time",              "10",                                               "Number",                                       False),
    ("Item Lead Time Period",       "DAYS",                                             "DAYS, WEEKS, or MONTHS",                       False),
    ("Item Payment Type",           "PER_INVOICE_ITEM",                                 "PER_INVOICE_ITEM or PER_DELIVERABLE",          False),
    ("Item Payment Terms",          '{"term": 2, "period": "MONTHS", "applied_from": "INVOICE_DATE"}', "JSON object",                  False),
    ("Item Deliverables Pmt Terms", '[{"allocation_percentage":"100","term":"2","period":"WEEKS","applied_from_milestone":"PO allocation date","applied_from_fixed_date":"2025-07-10"}]', "JSON array", False),
    ("Item Custom Sections",        "[]",                                               "JSON array",                                   False),

    # ── Delivery Schedule 1 ──
    ("Delivery Date 1",             "2026-06-30",                                       "YYYY-MM-DD",                                   False),
    ("Delivery Qty 1",              "10",                                               "Quantity for this delivery",                   False),
    ("Cost Centre 1",               "",                                                 "Cost centre UUID (optional)",                  False),
    ("General Ledger 1",            "",                                                 "GL UUID (optional)",                           False),
    ("Project 1",                   "",                                                 "Project UUID (optional)",                      False),

    # ── Delivery Schedule 2 (leave blank if only 1 delivery) ──
    ("Delivery Date 2",             "",                                                 "2nd delivery date — leave blank if only 1",    False),
    ("Delivery Qty 2",              "",                                                 "2nd delivery quantity",                        False),
    ("Cost Centre 2",               "",                                                 "",                                             False),
    ("General Ledger 2",            "",                                                 "",                                             False),
    ("Project 2",                   "",                                                 "",                                             False),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Upload"

    # Styles — optional columns
    header_font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill      = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hint_font        = Font(name="Calibri", size=9, italic=True, color="6B7280")
    hint_fill        = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    data_font        = Font(name="Calibri", size=11)
    # Styles — required columns (red header)
    req_header_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    req_header_fill  = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    req_hint_font    = Font(name="Calibri", size=9, italic=True, color="991B1B")
    req_hint_fill    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Row 1: headers, Row 2: hints, Row 3: example data
    for col_idx, (col_name, example, hint, required) in enumerate(COLUMNS, start=1):
        # Header
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = req_header_font if required else header_font
        cell.fill = req_header_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

        # Hint row
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = req_hint_font if required else hint_font
        cell.fill = req_hint_fill if required else hint_fill
        cell.alignment = Alignment(wrap_text=True)
        cell.border = thin_border

        # Example data row
        cell = ws.cell(row=3, column=col_idx, value=example)
        cell.font = data_font
        cell.border = thin_border

        # Auto-width (approx)
        max_len = max(len(col_name), len(hint), len(str(example)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLUMNS))}1"

    output = "PO BULK OPENAPI.xlsx"
    wb.save(output)
    print(f"Template saved: {output}")
    print(f"Columns: {len(COLUMNS)}")
    print(f"\nRow 1 = column headers  (red = required, blue = optional)")
    print(f"Row 2 = hints")
    print(f"Row 3 = example — fill your data from row 3 onwards")


if __name__ == "__main__":
    main()
